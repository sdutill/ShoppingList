import sys
import os
from os import path
import itertools
import openpyxl

# Locates the row index where a search term is the substring of a column value
def find_row_index(worksheet,search,col=1): # DONE
    index=None
    try:
        for row in range(1,worksheet.max_row+1):
            if search in worksheet.cell(column=col,row=_row).value:
                index=row
    except:
        pass
    if index != None:
        return index
    else:
        sys.exit(f'Row Indexing Error: {search} not found in any row of {worksheet.title}')


# Locates the column index where a search term is the substring of a column value
def find_column_index(worksheet,search,_row=1): # DONE
    index=None
    try:
        for col in range(1,worksheet.max_column+1):
            if search in worksheet.cell(column=col,row=_row).value:
                index=col
    except:
        pass
    if index != None:
        return index
    else:
        sys.exit(f'Column Indexing Error: {search} not found in any row of {worksheet.title}')


# Grabs the full column name based on the search term        
def find_column_header(worksheet,search,_row=1): # DONE
    column_header=None
    try:
        for col in range(1,worksheet.max_column+1):
            if search in worksheet.cell(column=col,row=_row).value:
                column_header=worksheet.cell(column=col,row=_row).value
    except:
        pass
    if column_header != None:
        return column_header
    else:
        sys.exit(f'Column Header Error: {search} not found in any row of {worksheet.title}')


#
def generate_key_set(sheet,key):  # DONE
    key_list={}
    key_index=find_column_index(sheet1)
    for row in range(1,worksheet.max_row+1):
        key_list.update({f'{sheet.cell(column=key_index,row=row).value}':{'Row':f'{col}'}})
    return set(key_list)


#
def generate_column_set(sheet,row_start=1): # DONE
    col_list={}
    for col in range(1,sheet.max_column+1):
        col_list.update({f'{sheet.cell(column=col,row=row_start).value}':{'Column':f'{col}'}})
    return col_list


#
def generate_set(sheet,key='',row_start=1): # DONE
    _list={}
    if key=='':
        for col in range(1,sheet.max_column+1):
            _list.update({f'{sheet.cell(column=col,row=row_start).value}':{'Column':f'{col}'}})
    else:
        key_index=find_column_index(sheet,key)
        for row in range(1,sheet.max_row+1):
            _list.update({f'{sheet.cell(column=key_index,row=row).value}':{'Row':f'{row}'}})
    return set(_list)


# Creates Dictionary of Excel spreadsheet; key value specified in function call; every column header accounted for
def build_dict(worksheet,key): # DONE
    key_list={}
    # Find Key Column
    key_column=find_column_index(worksheet=worksheet,search=key,_row=1)

    # Collect all the keys and their columns in the sheet
    for _row in range(2,worksheet.max_row+1):
        # Can't remember why I have this check
        if worksheet.cell(column=key_column,row=_row).value != None:
            # Container for row data
            row_data={}            
            # Iterate through the columns
            for col in range(1,worksheet.max_column+1):
                # Record this item's data in each column and include its original column name
                row_data.update({f'{worksheet.cell(column=col,row=1).value}':f'{worksheet.cell(column=col,row=_row).value}'})
            # Using the value of the key column for each row, add the row's data to the collection 
            key_list[f'{worksheet.cell(column=key_column,row=_row).value}']=row_data

    return key_list


